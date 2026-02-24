#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将测试结果CSV转换为Excel表格
按垂直角度x水平角度展示，使用颜色高亮显示
"""
import csv
import os
import glob
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    import sys
    sys.exit(1)

OUTPUT_PATH = r"D:\software\heiweilu\workspace\xgimi\code\20260206_dpl_auto"


def find_latest_result_file():
    """查找最新的测试结果文件"""
    pattern = os.path.join(OUTPUT_PATH, 'angle_test_result_*.csv')
    files = glob.glob(pattern)
    if not files:
        print("未找到测试结果文件！")
        return None
    latest_file = max(files, key=os.path.getmtime)
    return latest_file


def parse_coordinates(coord_str):
    """解析坐标字符串"""
    # 处理特殊情况：N/A或空字符串
    if not coord_str or coord_str.startswith('N/A'):
        return [(0, 0), (0, 0), (0, 0), (0, 0)]
    
    try:
        coords = [int(x.strip()) for x in coord_str.split(',')]
        return [
            (coords[0], coords[1]),  # Top Left
            (coords[2], coords[3]),  # Top Right
            (coords[4], coords[5]),  # Bottom Left
            (coords[6], coords[7])   # Bottom Right
        ]
    except (ValueError, IndexError) as e:
        print("Warning: Failed to parse coordinates '{}': {}".format(coord_str, e))
        return [(0, 0), (0, 0), (0, 0), (0, 0)]


def read_test_results(csv_file):
    """读取测试结果并组织数据"""
    data = {}
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 跳过预期失败的行（验证器标记为无效的）
            if row['Result'] == 'EXPECTED_FAIL':
                continue
            
            v_angle = int(row['VerticalAngle'])
            h_angle = int(row['HorizontalAngle'])
            
            # 兼容旧格式（TableCoords）和新格式（OriginalCoords/ClippedCoords）
            if 'ClippedCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)' in row:
                # 新格式：使用裁剪后的坐标
                table_coords = parse_coordinates(row['ClippedCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)'])
            elif 'TableCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)' in row:
                # 旧格式
                table_coords = parse_coordinates(row['TableCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)'])
            else:
                # 尝试OriginalCoords
                table_coords = parse_coordinates(row['OriginalCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)'])
            
            read_coords = parse_coordinates(row['ReadCoords(TL_x,TL_y,TR_x,TR_y,BL_x,BL_y,BR_x,BR_y)'])
            
            data[(v_angle, h_angle)] = {
                'angle_desc': row['AngleDesc'],
                'table_coords': table_coords,
                'read_coords': read_coords,
                'result': row['Result'],
                'error_code': int(row['ErrorCode']) if row['ErrorCode'] != 'N/A' else 0
            }
    
    return data


def generate_excel_table(data, output_file):
    """生成Excel表格"""
    
    # 自定义垂直角度排序：0, 10, 20, 30, 40, -10, -20, -30, -40
    v_angles_set = set(k[0] for k in data.keys())
    positive_v = sorted([a for a in v_angles_set if a > 0])
    negative_v = sorted([a for a in v_angles_set if a < 0], reverse=True)
    zero_v = [0] if 0 in v_angles_set else []
    v_angles = zero_v + positive_v + negative_v
    
    # 自定义水平角度排序：0, 10, 20, 30, -10, -20, -30
    h_angles_set = set(k[1] for k in data.keys())
    positive_angles = sorted([a for a in h_angles_set if a > 0])
    negative_angles = sorted([a for a in h_angles_set if a < 0], reverse=True)
    zero_angle = [0] if 0 in h_angles_set else []
    h_angles = zero_angle + positive_angles + negative_angles
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试结果"
    
    # 定义样式
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    v_header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    v_header_font = Font(bold=True, color="FFFFFF", size=10)
    
    angle_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    angle_font = Font(bold=True, color="1976D2", size=10)
    
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    warning_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 写入标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws['A1']
    cell.value = "垂直角度↓\n水平角度→"
    cell.fill = v_header_fill
    cell.font = v_header_font
    cell.alignment = center_alignment
    cell.border = border
    
    # 写入水平角度表头
    for col_idx, h_angle in enumerate(h_angles, start=2):
        if h_angle == 0:
            h_label = "0°"
        elif h_angle > 0:
            h_label = "上投{}°".format(h_angle)
        else:
            h_label = "下投{}°".format(abs(h_angle))
        
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h_label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 写入数据行
    current_row = 3
    for v_angle in v_angles:
        # 垂直角度标签
        if v_angle == 0:
            v_label = "0°"
        elif v_angle > 0:
            v_label = "左投{}°".format(v_angle)
        else:
            v_label = "右投{}°".format(abs(v_angle))
        
        cell = ws.cell(row=current_row, column=1)
        cell.value = v_label
        cell.fill = angle_fill
        cell.font = angle_font
        cell.alignment = center_alignment
        cell.border = border
        
        # 写入每个水平角度的数据
        for col_idx, h_angle in enumerate(h_angles, start=2):
            key = (v_angle, h_angle)
            cell = ws.cell(row=current_row, column=col_idx)
            
            if key in data:
                cell_data = data[key]
                is_pass = cell_data['result'] == 'PASS'
                
                # 设置单元格背景颜色
                if is_pass:
                    cell.fill = pass_fill
                elif cell_data['error_code'] == 1:
                    cell.fill = warning_fill
                else:
                    cell.fill = fail_fill
                
                # 构建单元格内容
                table_coords = cell_data['table_coords']
                read_coords = cell_data['read_coords']
                
                lines = []
                corner_names = ['TL', 'TR', 'BL', 'BR']
                
                for name, tc, rc in zip(corner_names, table_coords, read_coords):
                    coord_str = "{} ({},{})".format(name, tc[0], tc[1])
                    
                    # 如果有差异，显示最大差异
                    if tc != rc:
                        diff_x = abs(tc[0] - rc[0])
                        diff_y = abs(tc[1] - rc[1])
                        max_diff = max(diff_x, diff_y)
                        coord_str += " Δ{}".format(max_diff)
                    
                    lines.append(coord_str)
                
                # 添加结果标记
                if is_pass:
                    lines.append("✓")
                else:
                    if cell_data['error_code'] == 1:
                        lines.append("✗ EC:1")
                    else:
                        lines.append("✗ EC:{}".format(cell_data['error_code']))
                
                cell.value = "\n".join(lines)
            else:
                cell.value = "-"
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            cell.alignment = center_alignment
            cell.border = border
        
        current_row += 1
    
    # 设置列宽和行高
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, len(h_angles) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    for row_idx in range(3, current_row):
        ws.row_dimensions[row_idx].height = 80
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    
    # 添加图例说明（在表格下方）
    legend_row = current_row + 2
    legend_end_col = len(h_angles) + 1
    
    if legend_end_col > 1:  # 确保至少有2列才合并
        ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=legend_end_col)
    
    legend_cell = ws.cell(row=legend_row, column=1)
    legend_cell.value = ("说明：\n"
                        "绿色=PASS(测试通过，坐标完全匹配) | "
                        "蓝色=FAIL但EC=1(硬件执行成功但坐标有偏差) | "
                        "红色=FAIL且EC≠1(硬件执行失败)\n"
                        "TL/TR/BL/BR=左上/右上/左下/右下角点 | "
                        "Δ=读取与写入坐标的最大差异(像素) | "
                        "EC=ErrorCode")
    legend_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    legend_cell.font = Font(size=9, color="666666")
    ws.row_dimensions[legend_row].height = 50
    
    # 保存文件
    wb.save(output_file)
    print("Excel表格已生成: {}".format(output_file))


def main():
    """主函数"""
    print("=" * 80)
    print("生成Excel测试结果表格")
    print("=" * 80)
    
    # 查找最新的结果文件
    result_file = find_latest_result_file()
    if not result_file:
        print("请先运行角度测试脚本生成结果文件！")
        return
    
    print("\n读取结果文件: {}".format(os.path.basename(result_file)))
    
    # 读取测试结果
    data = read_test_results(result_file)
    print("读取到 {} 组测试数据\n".format(len(data)))
    
    # 生成Excel表格
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    excel_file = os.path.join(OUTPUT_PATH, 'test_result_table_{}.xlsx'.format(timestamp))
    
    print("生成Excel表格...")
    generate_excel_table(data, excel_file)
    
    print("\n" + "=" * 80)
    print("完成！")
    print("=" * 80)
    print("\n文件位置:")
    print("  {}".format(excel_file))
    print("\n可以直接使用Excel打开，或上传到在线表格（腾讯文档、Google Sheets等）")
    print("=" * 80)


if __name__ == "__main__":
    main()
