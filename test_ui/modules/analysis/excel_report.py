# -*- coding: utf-8 -*-
"""
Excel 报表生成模块

原始脚本: 202602027_dlp_auto/src/Analysis/Trapezoidal_angle_accuracy_output_in_Excel.py
功能: 读取角度测试结果CSV，生成格式化Excel报表（行=Yaw，列=Pitch，颜色标注PASS/FAIL）
"""

import os
import csv
import glob
import pandas as pd
from datetime import datetime

MODULE_INFO = {
    "name": "Excel报表生成",
    "category": "analysis",
    "description": "将角度测试结果转换为格式化Excel报表。\n"
                   "行=Yaw角度，列=Pitch角度，单元格颜色标注PASS(绿)/FAIL(红)。",
    "input_type": "csv",
    "input_description": "角度测试结果CSV，需含 Yaw, Pitch, Result, WriteCoords 列。\n"
                         "留空则自动搜索最新结果文件。",
    "output_type": "excel",
    "script_file": "excel_report.py",
    "reference_output_desc": "输出Excel报表，包含角度测试汇总表、Yaw/Pitch各登山数据PASS分布、Delta差异分析、各模块单元格式化结果。",
    "params": [],
}


def run(input_path: str, output_dir: str, params: dict,
        progress_callback=None, log_callback=None) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _log(msg, level='INFO'):
        if log_callback:
            log_callback(msg, level)

    def _progress(cur, total):
        if progress_callback:
            progress_callback(cur, total)

    try:
        # 如果没有指定文件，尝试自动查找
        if not input_path or not os.path.exists(input_path):
            project_root = params.get('project_root', '')
            if project_root:
                from core.file_utils import find_latest_file
                search_dir = os.path.join(project_root, 'reports', 'Angle_test_results')
                input_path = find_latest_file(search_dir, 'angle_test_result_*.csv')
                if input_path:
                    _log(f"自动找到最新结果: {input_path}")
                else:
                    return {"status": "error", "message": "未找到角度测试结果文件"}
            else:
                return {"status": "error", "message": "请选择输入文件"}

        _log(f"加载: {input_path}")
        _progress(1, 10)

        from core.data_loader import load_angle_test_result
        from core.coord_parser import parse_as_tuples

        df = load_angle_test_result(input_path, log_callback=log_callback)
        for col in ['Yaw', 'Pitch', 'Result']:
            if col not in df.columns:
                return {"status": "error", "message": f"缺少列: {col}"}
        _progress(3, 10)

        # 提取坐标（如果有）
        has_coords = 'WriteCoords' in df.columns

        # 组织数据: {(yaw, pitch): {'result': ..., 'coords': ...}}
        data = {}
        for _, row in df.iterrows():
            yaw = round(float(row['Yaw']), 1)
            pitch = round(float(row['Pitch']), 1)
            result = str(row['Result']).strip().upper()
            coords_str = str(row['WriteCoords']) if has_coords else ''
            data[(yaw, pitch)] = {
                'result': result,
                'coords': coords_str,
            }
        _progress(5, 10)

        yaw_values = sorted(set(k[0] for k in data.keys()))
        pitch_values = sorted(set(k[1] for k in data.keys()))
        _log(f"Yaw: {len(yaw_values)} 个值, Pitch: {len(pitch_values)} 个值")

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '角度测试结果'

        # 样式
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font_white = Font(bold=True, size=10, color='FFFFFF')
        pass_fill = PatternFill('solid', fgColor='C6EFCE')
        fail_fill = PatternFill('solid', fgColor='FFC7CE')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 表头
        ws.cell(row=1, column=1, value='Yaw \\ Pitch')
        ws.cell(row=1, column=1).font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align

        for ci, pitch in enumerate(pitch_values):
            cell = ws.cell(row=1, column=ci+2, value=f'{pitch:.1f}°')
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
        _progress(6, 10)

        # 数据行
        for ri, yaw in enumerate(yaw_values):
            # Yaw 标题
            yaw_cell = ws.cell(row=ri+2, column=1, value=f'{yaw:.1f}°')
            yaw_cell.font = header_font
            yaw_cell.alignment = center_align

            for ci, pitch in enumerate(pitch_values):
                entry = data.get((yaw, pitch))
                cell = ws.cell(row=ri+2, column=ci+2)

                if entry:
                    result = entry['result']
                    cell.value = result
                    cell.fill = pass_fill if result == 'PASS' else fail_fill
                else:
                    cell.value = '-'

                cell.alignment = center_align
                cell.border = thin_border

            _progress(6 + int(ri / max(len(yaw_values), 1) * 3), 10)

        # 自动列宽
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 9
        ws.column_dimensions['A'].width = 12

        # 保存
        from core.file_utils import make_output_path
        project_root = params.get('project_root', output_dir)
        _, output_path = make_output_path(
            project_root, 'Angle_test_results', '',
            prefix='test_result_table', ext='.xlsx')
        wb.save(output_path)
        _log(f"Excel已保存: {output_path}", "SUCCESS")
        _log(f"矩阵大小: {len(yaw_values)} × {len(pitch_values)}")
        _progress(10, 10)

        return {"status": "success", "output_path": output_path, "figure": None,
                "message": f"Excel {len(yaw_values)}×{len(pitch_values)} 矩阵"}

    except Exception as e:
        import traceback
        return {"status": "error", "message": f"{e}\n{traceback.format_exc()}"}
